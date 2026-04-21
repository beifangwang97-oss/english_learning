import hashlib
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor
from queue import Empty, Queue

import pandas as pd
import streamlit as st
from openai import OpenAI
from openpyxl import load_workbook


st.set_page_config(
    page_title="Tiger English - Phrase Merge Tool",
    layout="wide",
    page_icon="T",
)

st.markdown(
    """
<style>
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding-left: 20px;
        padding-right: 20px;
        font-weight: 500;
    }
    .stTabs [aria-selected="true"] {
        background-color: #1f77b4;
        color: white;
    }
    div[data-testid="stVerticalBlock"] > div[style*="flex-direction: column"] > div > div > div > div {
        opacity: 1 !important;
    }
</style>
""",
    unsafe_allow_html=True,
)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
PENDING_PHRASE_DIR = os.path.join(BASE_DIR, "待处理短语")
DEFAULT_BASE_URL = "https://openrouter.ai/api/v1"
DEFAULT_MODEL = "openai/gpt-4o-mini"
LLM_CLIENT_CACHE = {}
MAX_RETRIES = 3

OPENROUTER_MODEL_OPTIONS = [
    {"id": "openai/gpt-4o-mini", "name": "OpenAI GPT-4o mini", "input_price": "$0.15/M", "output_price": "$0.60/M"},
    {"id": "openai/gpt-4.1-mini", "name": "OpenAI GPT-4.1 mini", "input_price": "$0.40/M", "output_price": "$1.60/M"},
    {"id": "openai/gpt-4.1", "name": "OpenAI GPT-4.1", "input_price": "$2.00/M", "output_price": "$8.00/M"},
    {"id": "openai/gpt-4o", "name": "OpenAI GPT-4o", "input_price": "$2.50/M", "output_price": "$10.00/M"},
    {"id": "google/gemini-2.5-flash-lite", "name": "Google Gemini 2.5 Flash Lite", "input_price": "$0.10/M", "output_price": "$0.40/M"},
    {"id": "google/gemini-2.5-flash", "name": "Google Gemini 2.5 Flash", "input_price": "$0.15/M", "output_price": "$0.60/M"},
    {"id": "google/gemini-2.5-pro", "name": "Google Gemini 2.5 Pro", "input_price": "$1.25/M", "output_price": "$10.00/M"},
    {
        "id": "qwen/qwen-2.5-vl-72b-instruct",
        "name": "Qwen 2.5 VL 72B Instruct",
        "input_price": "OpenRouter实时价",
        "output_price": "OpenRouter实时价",
    },
]


def sanitize_text(value):
    return str(value or "").strip()


def rel_path(path):
    return os.path.relpath(path, BASE_DIR).replace("\\", "/")


def ensure_parent_dir(path):
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)


def _model_label(option):
    return (
        f"{option['name']} | {option['id']} | "
        f"输入 {option['input_price']} 输出 {option['output_price']}"
    )


OPENROUTER_MODEL_LABEL_TO_ID = {_model_label(opt): opt["id"] for opt in OPENROUTER_MODEL_OPTIONS}
OPENROUTER_MODEL_ID_TO_LABEL = {opt["id"]: _model_label(opt) for opt in OPENROUTER_MODEL_OPTIONS}
OPENROUTER_MODEL_LABELS = list(OPENROUTER_MODEL_LABEL_TO_ID.keys())


def render_model_selector(widget_label, default_model_id, picker_key):
    safe_default_id = default_model_id if default_model_id in OPENROUTER_MODEL_ID_TO_LABEL else DEFAULT_MODEL
    default_label = OPENROUTER_MODEL_ID_TO_LABEL[safe_default_id]
    selected = st.multiselect(
        widget_label,
        options=OPENROUTER_MODEL_LABELS,
        default=[default_label],
        max_selections=1,
        key=picker_key,
    )
    if selected:
        return OPENROUTER_MODEL_LABEL_TO_ID[selected[0]]
    return safe_default_id


def get_openai_client(api_key, base_url):
    cache_key = f"{base_url}|{api_key}"
    if cache_key not in LLM_CLIENT_CACHE:
        LLM_CLIENT_CACHE[cache_key] = OpenAI(api_key=api_key, base_url=base_url)
    return LLM_CLIENT_CACHE[cache_key]


def pick_api_key_for_job(api_keys, job_idx):
    if not api_keys:
        return ""
    return api_keys[job_idx % len(api_keys)]


def build_api_slot_label(api_keys, job_idx):
    if not api_keys:
        return "API-1"
    return f"API-{(job_idx % len(api_keys)) + 1}"


def build_copy_output_path(source_path):
    root, ext = os.path.splitext(source_path)
    candidate = f"{root}_mode8{ext}"
    if not os.path.exists(candidate):
        return candidate
    index = 2
    while True:
        candidate = f"{root}_mode8_{index}{ext}"
        if not os.path.exists(candidate):
            return candidate
        index += 1


def normalize_grade_semester_stem(stem):
    value = sanitize_text(stem).replace("_", "").replace(" ", "")
    match = re.match(r"^(.*?)(上册|下册|全一册|全册)$", value)
    if not match:
        return "", ""
    return sanitize_text(match.group(1)), sanitize_text(match.group(2))


def build_grade_semester_label(grade, semester):
    return f"{sanitize_text(grade)}_{sanitize_text(semester)}"


def normalize_unit_label(unit_text):
    text = sanitize_text(unit_text)
    if not text:
        return ""
    match = re.match(r"(?i)^unit\s*([0-9]+)$", text)
    if match:
        return f"Unit {match.group(1)}"
    match = re.match(r"(?i)^starter\s*unit\s*([0-9]+)$", text)
    if match:
        return f"Starter Unit {match.group(1)}"
    return text


def normalize_phrase_key(text):
    value = sanitize_text(text)
    value = value.replace("\u3000", " ")
    value = value.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
    value = re.sub(r"\s+", " ", value)
    return value.casefold()


def sanitize_phrase_display(text):
    value = sanitize_text(text)
    value = value.replace("\u3000", " ")
    value = re.sub(r"\s+", " ", value)
    return value


def infer_source_tag_from_filename(filename):
    stem = os.path.splitext(os.path.basename(str(filename or "")))[0]
    if stem.endswith("_current"):
        return "current_book"
    if stem.endswith("_primary"):
        return "primary_school_review"
    return ""


def phrase_unique_id(word, unit, grade, semester="", book_version="", source_tag=""):
    raw_str = f"{word}_{unit}_{grade}_{semester}_{book_version}_{source_tag}"
    return hashlib.md5(raw_str.encode("utf-8")).hexdigest()[:10]


def discover_pending_phrase_excels():
    results = []
    if not os.path.isdir(PENDING_PHRASE_DIR):
        return results
    for root, _, files in os.walk(PENDING_PHRASE_DIR):
        for name in files:
            if not name.lower().endswith(".xlsx"):
                continue
            abs_path = os.path.join(root, name)
            rel_name = os.path.relpath(abs_path, PENDING_PHRASE_DIR).replace("\\", "/")
            rel_dir = os.path.relpath(root, PENDING_PHRASE_DIR).replace("\\", "/")
            book_version = "" if rel_dir == "." else sanitize_text(rel_dir.split("/")[0])
            grade, semester = normalize_grade_semester_stem(os.path.splitext(name)[0])
            results.append(
                {
                    "abs_path": abs_path,
                    "raw_name": name,
                    "display_name": rel_name,
                    "book_version": book_version,
                    "grade": grade,
                    "semester": semester,
                }
            )
    return sorted(results, key=lambda item: item["display_name"])


def discover_target_phrase_jsonls(book_version, grade, semester):
    folder = os.path.join(DATA_DIR, sanitize_text(book_version), "短语")
    if not os.path.isdir(folder):
        return []
    prefix = build_grade_semester_label(grade, semester)
    results = []
    for name in os.listdir(folder):
        if not name.lower().endswith(".jsonl"):
            continue
        if not sanitize_text(name).startswith(prefix):
            continue
        results.append(os.path.join(folder, name))
    return sorted(results)


def load_jsonl_rows(path):
    rows = []
    if not os.path.exists(path):
        return rows
    with open(path, "r", encoding="utf-8-sig") as rf:
        for line in rf:
            if not line.strip():
                continue
            rows.append(json.loads(line))
    return rows


def write_jsonl_rows(path, rows):
    ensure_parent_dir(path)
    with open(path, "w", encoding="utf-8") as wf:
        for row in rows:
            wf.write(json.dumps(row, ensure_ascii=False) + "\n")


def read_phrase_excel_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [sanitize_text(cell) for cell in (header_row or [])]
    col_map = {header: idx for idx, header in enumerate(headers)}
    phrase_idx = col_map.get("英文单词", 0)
    meaning_idx = col_map.get("中文意思", 2)
    unit_idx = col_map.get("单元", 5)
    rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row)
        phrase = sanitize_phrase_display(values[phrase_idx] if phrase_idx < len(values) else "")
        meaning = sanitize_text(values[meaning_idx] if meaning_idx < len(values) else "")
        unit = normalize_unit_label(values[unit_idx] if unit_idx < len(values) else "")
        if not phrase:
            continue
        rows.append(
            {
                "excel_row_no": row_no,
                "word": phrase,
                "meaning": meaning,
                "unit": unit,
            }
        )
    return rows


def count_pending_phrase_candidates(excel_file, target_file):
    source_rows = load_jsonl_rows(target_file)
    existing_keys = {normalize_phrase_key(row.get("word", "")) for row in source_rows if isinstance(row, dict)}
    count = 0
    for item in read_phrase_excel_rows(excel_file):
        phrase_key = normalize_phrase_key(item["word"])
        if not phrase_key or phrase_key in existing_keys:
            continue
        existing_keys.add(phrase_key)
        count += 1
    return count


def build_prompt_payload(items):
    return json.dumps(items, ensure_ascii=False)


def extract_json_object(text):
    content = sanitize_text(text)
    if not content:
        return {}
    try:
        return json.loads(content)
    except Exception:
        match = re.search(r"\{.*\}", content, flags=re.S)
        if match:
            return json.loads(match.group(0))
    return {}


def validate_example_result(item):
    example = sanitize_text(item.get("example"))
    example_zh = sanitize_text(item.get("example_zh"))
    if not example or not example_zh:
        return {"example": "", "example_zh": ""}, "empty_result"
    if not re.search(r"[A-Za-z]", example):
        return {"example": "", "example_zh": ""}, "invalid_example"
    return {"example": example, "example_zh": example_zh}, "valid"


def call_example_fill_api(batch_items, api_key, base_url, model_name):
    client = get_openai_client(api_key, base_url)
    payload = build_prompt_payload(batch_items)
    prompt = f"""
你是英语教材短语例句补全助手。请根据短语、中文释义、单元信息，为每条短语生成适合中学生学习的英文例句和对应中文翻译。

要求：
1. 只输出 JSON 对象。
2. 返回格式必须是：
{{
  "items": [
    {{
      "line_no": 1,
      "example": "She is from Beijing.",
      "example_zh": "她来自北京。"
    }}
  ]
}}
3. 必须按输入 line_no 原样返回。
4. 例句要自然、简洁、适合教材，不要过长。
5. 如果短语中带有 sb. / sth. / sp. / one's 等占位符，请在例句里替换成自然具体的内容。
6. 中文翻译要与英文例句一一对应。
7. 如果没有把握，请把 example 和 example_zh 都置空。

输入数据：
{payload}
""".strip()

    response = client.chat.completions.create(
        model=model_name,
        messages=[
            {"role": "system", "content": "你只返回合法 JSON，不要输出额外说明。"},
            {"role": "user", "content": prompt},
        ],
        temperature=0.2,
        response_format={"type": "json_object"},
    )
    content = response.choices[0].message.content or "{}"
    parsed = extract_json_object(content)
    items = parsed.get("items", [])
    if not isinstance(items, list):
        return []
    return items


def enrich_examples_with_retry(batch_items, api_key, base_url, model_name):
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            raw_items = call_example_fill_api(batch_items, api_key, base_url, model_name)
            result_map = {}
            reason_map = {}
            for raw in raw_items:
                if not isinstance(raw, dict):
                    continue
                line_no = raw.get("line_no")
                try:
                    line_no = int(line_no)
                except Exception:
                    continue
                payload, reason = validate_example_result(raw)
                result_map[line_no] = payload
                reason_map[line_no] = reason
            return result_map, reason_map
        except Exception as exc:
            last_error = exc
            time.sleep(attempt)
    raise RuntimeError(f"批次补全失败：{last_error}")


def process_phrase_job(job, api_key, api_label, base_url, model_name, batch_size, event_queue):
    source_rows = load_jsonl_rows(job["target_file"])
    existing_keys = {normalize_phrase_key(row.get("word", "")) for row in source_rows if isinstance(row, dict)}
    excel_rows = read_phrase_excel_rows(job["excel_file"])
    pending_items = []
    skipped_duplicates = 0

    for idx, item in enumerate(excel_rows, start=1):
        phrase_key = normalize_phrase_key(item["word"])
        if not phrase_key:
            continue
        if phrase_key in existing_keys:
            skipped_duplicates += 1
            continue
        existing_keys.add(phrase_key)
        pending_items.append(
            {
                "line_no": idx,
                "word": item["word"],
                "meaning": item["meaning"],
                "unit": item["unit"],
            }
        )

    rows = list(source_rows)
    if not pending_items:
        write_jsonl_rows(job["output_file"], rows)
        event_queue.put(
            {
                "type": "job_done",
                "uid": job["uid"],
                "inserted_rows": 0,
                "skipped_duplicates": skipped_duplicates,
                "failed_rows": 0,
            }
        )
        return {
            "source_excel": job["excel_file"],
            "target_file": job["target_file"],
            "output_file": job["output_file"],
            "excel_rows": len(excel_rows),
            "inserted_rows": 0,
            "skipped_duplicates": skipped_duplicates,
            "failed_rows": 0,
            "finished_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }

    chunks = [pending_items[i : i + batch_size] for i in range(0, len(pending_items), batch_size)]
    event_queue.put(
        {
            "type": "job_started",
            "uid": job["uid"],
            "api_label": api_label,
            "total_batches": len(chunks),
            "total_candidates": len(pending_items),
            "skipped_duplicates": skipped_duplicates,
        }
    )

    inserted_rows = 0
    failed_rows = 0
    debug_reason_map = {
        "valid": "例句与译文已生成并写入。",
        "empty_result": "模型未返回有效例句，已按空例句写入。",
        "invalid_example": "模型返回的英文例句无效，已按空例句写入。",
        "batch_error": "批次请求失败，已按空例句写入。",
    }

    for chunk_idx, chunk in enumerate(chunks, start=1):
        result_map = {}
        reason_map = {}
        batch_error = ""
        try:
            result_map, reason_map = enrich_examples_with_retry(chunk, api_key, base_url, model_name)
        except Exception as exc:
            batch_error = str(exc)

        for item in chunk:
            payload = result_map.get(item["line_no"], {"example": "", "example_zh": ""})
            reason = reason_map.get(item["line_no"], "empty_result")
            if batch_error:
                payload = {"example": "", "example_zh": ""}
                reason = "batch_error"
                failed_rows += 1
            elif not payload.get("example") or not payload.get("example_zh"):
                failed_rows += 1

            new_row = {
                "word": item["word"],
                "phonetic": "",
                "meanings": [
                    {
                        "pos": "",
                        "meaning": item["meaning"],
                        "example": payload.get("example", ""),
                        "example_zh": payload.get("example_zh", ""),
                    }
                ],
                "unit": item["unit"],
                "type": "phrase",
                "book_version": job["book_version"],
                "grade": job["grade"],
                "semester": job["semester"],
                "id": phrase_unique_id(
                    item["word"],
                    item["unit"],
                    job["grade"],
                    job["semester"],
                    job["book_version"],
                    job["source_tag"],
                ),
                "phrase_audio": "",
            }
            if job["source_tag"]:
                new_row["source_tag"] = job["source_tag"]

            rows.append(new_row)
            inserted_rows += 1

            event_queue.put(
                {
                    "type": "item",
                    "uid": job["uid"],
                    "row": {
                        "教材": job["job_label"],
                        "单元": item["unit"],
                        "短语": item["word"],
                        "释义": item["meaning"],
                        "例句": payload.get("example", ""),
                        "译文": payload.get("example_zh", ""),
                    },
                }
            )
            event_queue.put(
                {
                    "type": "debug",
                    "uid": job["uid"],
                    "row": {
                        "教材": job["job_label"],
                        "短语": item["word"],
                        "API": api_label,
                        "状态": "已写入" if reason == "valid" else "已校正",
                        "说明": debug_reason_map.get(reason, "结果已处理。"),
                    },
                }
            )

        write_jsonl_rows(job["output_file"], rows)
        event_queue.put(
            {
                "type": "batch_done",
                "uid": job["uid"],
                "batch_idx": chunk_idx,
                "total_batches": len(chunks),
                "inserted_rows": inserted_rows,
                "failed_rows": failed_rows,
                "last_error": batch_error,
            }
        )

    event_queue.put(
        {
            "type": "job_done",
            "uid": job["uid"],
            "inserted_rows": inserted_rows,
            "skipped_duplicates": skipped_duplicates,
            "failed_rows": failed_rows,
        }
    )
    return {
        "source_excel": job["excel_file"],
        "target_file": job["target_file"],
        "output_file": job["output_file"],
        "excel_rows": len(excel_rows),
        "inserted_rows": inserted_rows,
        "skipped_duplicates": skipped_duplicates,
        "failed_rows": failed_rows,
        "finished_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }


def render_summary(summary_rows):
    if not summary_rows:
        return
    table = pd.DataFrame(
        [
            {
                "源Excel": rel_path(item["source_excel"]),
                "目标JSONL": rel_path(item["target_file"]),
                "输出JSONL": rel_path(item["output_file"]),
                "Excel行数": item["excel_rows"],
                "新增短语": item["inserted_rows"],
                "重复跳过": item["skipped_duplicates"],
                "空例句行": item["failed_rows"],
                "完成时间": item["finished_at"],
            }
            for item in summary_rows
        ]
    )
    st.dataframe(table, use_container_width=True, hide_index=True)


def main():
    with st.sidebar:
        st.header("补全 API 配置")
        base_url = st.text_input("补全 Base URL", value=DEFAULT_BASE_URL, key="mode8_base_url")
        model_name = render_model_selector(
            "补全模型（下拉勾选，单选）",
            st.session_state.get("mode8_model_name", DEFAULT_MODEL),
            "mode8_model_picker",
        )
        st.session_state.mode8_model_name = model_name
        st.caption(f"补全模型ID：`{model_name}`")

        if "mode8_api_key_count" not in st.session_state:
            st.session_state.mode8_api_key_count = 1

        key_col, add_col = st.columns([6, 1])
        with key_col:
            st.text_input("补全 API Key", type="password", key="mode8_api_key_1")
        with add_col:
            st.write("")
            st.write("")
            if st.button("＋", key="btn_mode8_add_api", help="新增一个 API Key 输入框"):
                st.session_state.mode8_api_key_count += 1
                st.rerun()

        for idx in range(2, int(st.session_state.mode8_api_key_count) + 1):
            row_col, del_col = st.columns([6, 1])
            with row_col:
                st.text_input(f"补全 API Key {idx}", type="password", key=f"mode8_api_key_{idx}")
            with del_col:
                st.write("")
                st.write("")
                if st.button("－", key=f"btn_mode8_del_api_{idx}", help=f"删除 API Key {idx}"):
                    total = int(st.session_state.mode8_api_key_count)
                    for j in range(idx, total):
                        st.session_state[f"mode8_api_key_{j}"] = st.session_state.get(f"mode8_api_key_{j+1}", "")
                    st.session_state.pop(f"mode8_api_key_{total}", None)
                    st.session_state.mode8_api_key_count = max(1, total - 1)
                    st.rerun()

        api_keys = []
        for idx in range(1, int(st.session_state.mode8_api_key_count) + 1):
            value = sanitize_text(st.session_state.get(f"mode8_api_key_{idx}", ""))
            if value and value not in api_keys:
                api_keys.append(value)

        st.caption(f"当前可用补全 Key 数量：{len(api_keys)}")
        if st.button("一键测试全部 API", key="btn_mode8_test_all_api"):
            if not api_keys:
                st.warning("请先输入至少一个 API Key。")
            else:
                test_rows = []
                with st.spinner("正在测试 API 连通性..."):
                    for i, api_key in enumerate(api_keys, start=1):
                        ok = False
                        msg = ""
                        try:
                            client = get_openai_client(api_key, base_url)
                            _ = client.chat.completions.create(
                                model=model_name,
                                messages=[{"role": "user", "content": "ping"}],
                                temperature=0,
                                max_tokens=1,
                                timeout=20,
                            )
                            ok = True
                            msg = f"第{i}个API可用"
                        except Exception as exc:
                            msg = f"第{i}个API不可用：{exc}"
                        test_rows.append((ok, msg))

                for ok, msg in test_rows:
                    if ok:
                        st.success(msg)
                    else:
                        st.error(msg)
        output_mode = st.radio(
            "输出方式",
            options=["原文件写回", "同目录新建副本"],
            horizontal=False,
            key="mode8_output_mode",
        )
        st.caption("选择“同目录新建副本”时，会在目标短语 JSONL 同目录生成 `_mode8.jsonl`，如重名则自动顺延为 `_mode8_2.jsonl`、`_mode8_3.jsonl`。")
        batch_size = st.number_input("每批次短语数", min_value=5, max_value=50, value=15, step=5, key="mode8_batch_size")

    tab1 = st.tabs(["模式八：短语补充与标准化"])[0]
    with tab1:
        st.subheader("模式八：短语 Excel 补充导入（去重合并 + API 补例句）")
        st.caption("扫描 `待处理短语/版本/年级册数.xlsx`，匹配现有短语 JSONL，按短语英文去重后插入缺失短语，并调用 API 仅补全例句与例句翻译。")

        excel_items = discover_pending_phrase_excels()
        excel_map = {item["display_name"]: item for item in excel_items}
        select_all = st.checkbox("默认全选当前扫描到的短语 Excel", value=False, key="mode8_select_all")
        selected_labels = st.multiselect(
            "从本地目录选择待处理短语 Excel",
            options=list(excel_map.keys()),
            default=list(excel_map.keys()) if select_all else [],
            key="mode8_excel_selector",
        )

        matched_jobs = []
        unmatched_labels = []
        invalid_meta_labels = []

        if selected_labels:
            for idx, label in enumerate(selected_labels):
                item = excel_map[label]
                if not item["book_version"] or not item["grade"] or not item["semester"]:
                    invalid_meta_labels.append(label)
                    continue
                targets = discover_target_phrase_jsonls(item["book_version"], item["grade"], item["semester"])
                if not targets:
                    unmatched_labels.append(label)
                    continue
                for target_idx, target_file in enumerate(targets, start=1):
                    source_tag = infer_source_tag_from_filename(os.path.basename(target_file))
                    output_file = target_file if output_mode == "原文件写回" else build_copy_output_path(target_file)
                    pending_count = count_pending_phrase_candidates(item["abs_path"], target_file)
                    uid = hashlib.md5(f"{idx}|{target_idx}|{item['abs_path']}|{target_file}".encode("utf-8")).hexdigest()[:12]
                    matched_jobs.append(
                        {
                            "uid": uid,
                            "job_label": f"{label} -> {os.path.basename(target_file)}",
                            "excel_file": item["abs_path"],
                            "target_file": target_file,
                            "output_file": output_file,
                            "book_version": item["book_version"],
                            "grade": item["grade"],
                            "semester": item["semester"],
                            "source_tag": source_tag,
                            "pending_count": pending_count,
                        }
                    )

            if matched_jobs:
                st.success(f"匹配成功：{len(matched_jobs)} 个任务")
                for job in matched_jobs:
                    st.write(f"- {job['job_label']}")
            if unmatched_labels:
                st.warning("以下 Excel 未匹配到目标短语 JSONL：")
                for label in unmatched_labels:
                    st.write(f"- {label}")
            if invalid_meta_labels:
                st.warning("以下 Excel 文件名或目录结构无法识别版本/年级/册数，已跳过：")
                for label in invalid_meta_labels:
                    st.write(f"- {label}")
        else:
            st.info("请先从本地目录选择一个或多个短语 Excel。")

        if "mode8_last_summary" not in st.session_state:
            st.session_state.mode8_last_summary = []

        c_start, c_info = st.columns([3, 2])
        with c_start:
            run_button = st.button(
                "开始补充并写入（模式八）",
                type="primary",
                disabled=(not matched_jobs or not api_keys),
                key="btn_mode8_run",
            )
        with c_info:
            st.caption("处理过程中会按任务批次持续写出到目标短语 JSONL。")

        status_placeholder = st.empty()
        mode8_progress = st.empty()
        mode8_table_placeholder = st.empty()
        mode8_detail_placeholder = st.empty()

        if run_button:
            live_rows = {}
            live_recent = {}
            live_debug = {}
            detail_placeholders = {}
            debug_placeholders = {}
            summary_rows = []
            event_queue = Queue()

            with mode8_detail_placeholder.container():
                st.markdown("**实时回传（按教材分组，逐条追加）**")
                for job_idx, job in enumerate(matched_jobs):
                    api_label = build_api_slot_label(api_keys, job_idx)
                    with st.expander(job["job_label"], expanded=False):
                        st.caption(f"输出：`{rel_path(job['output_file'])}` | 分配 {api_label}")
                        with st.expander("结果明细", expanded=True):
                            detail_placeholders[job["uid"]] = st.empty()
                        with st.expander("调试日志", expanded=False):
                            debug_placeholders[job["uid"]] = st.empty()

            total_batches = 0
            for job_idx, job in enumerate(matched_jobs):
                total_excel_rows = len(read_phrase_excel_rows(job["excel_file"]))
                total_candidate_batches = max(1, (job["pending_count"] + int(batch_size) - 1) // int(batch_size)) if job["pending_count"] else 1
                total_batches += total_candidate_batches
                live_rows[job["uid"]] = {
                    "教材": job["job_label"],
                    "状态": "queued",
                    "API序号": build_api_slot_label(api_keys, job_idx),
                    "总批次": total_candidate_batches,
                    "已完成": 0,
                    "Excel短语": total_excel_rows,
                    "待新增": job["pending_count"],
                    "新增写入": 0,
                    "重复跳过": 0,
                    "空例句行": 0,
                    "输出文件": rel_path(job["output_file"]),
                }
                live_recent[job["uid"]] = []
                live_debug[job["uid"]] = []

            def render_live():
                with mode8_table_placeholder.container():
                    st.markdown("**提取进度（按教材）**")
                    st.dataframe(pd.DataFrame(list(live_rows.values())), use_container_width=True, hide_index=True)

            def push_recent(uid, row):
                rows = live_recent.setdefault(uid, [])
                rows.append(row)
                if len(rows) > 200:
                    live_recent[uid] = rows[-200:]
                    rows = live_recent[uid]
                ph = detail_placeholders.get(uid)
                if ph:
                    ph.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            def push_debug(uid, row):
                rows = live_debug.setdefault(uid, [])
                rows.append(row)
                if len(rows) > 300:
                    live_debug[uid] = rows[-300:]
                    rows = live_debug[uid]
                ph = debug_placeholders.get(uid)
                if ph:
                    ph.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            def worker(job_idx, job):
                return process_phrase_job(
                    job=job,
                    api_key=pick_api_key_for_job(api_keys, job_idx),
                    api_label=build_api_slot_label(api_keys, job_idx),
                    base_url=base_url,
                    model_name=model_name,
                    batch_size=int(batch_size),
                    event_queue=event_queue,
                )

            max_workers = max(1, min(len(api_keys), len(matched_jobs)))
            done_batches = 0
            render_live()
            mode8_progress.progress(0.0)
            status_placeholder.info(f"开始并行处理：任务 {len(matched_jobs)} 个，API {len(api_keys)} 个，并发 {max_workers}。")

            executor = ThreadPoolExecutor(max_workers=max_workers)
            futures = {executor.submit(worker, job_idx, job): job for job_idx, job in enumerate(matched_jobs)}

            while True:
                had_event = False
                while True:
                    try:
                        event = event_queue.get_nowait()
                    except Empty:
                        break
                    had_event = True
                    uid = event.get("uid")
                    row_ref = live_rows.get(uid)
                    if not row_ref:
                        continue
                    event_type = event.get("type")
                    if event_type == "job_started":
                        row_ref["状态"] = "running"
                        row_ref["重复跳过"] = int(event.get("skipped_duplicates", row_ref.get("重复跳过", 0)))
                    elif event_type == "item":
                        push_recent(uid, event.get("row", {}))
                    elif event_type == "debug":
                        push_debug(uid, event.get("row", {}))
                    elif event_type == "batch_done":
                        row_ref["状态"] = "running"
                        row_ref["已完成"] = int(event.get("batch_idx", row_ref.get("已完成", 0)))
                        row_ref["新增写入"] = int(event.get("inserted_rows", row_ref.get("新增写入", 0)))
                        row_ref["空例句行"] = int(event.get("failed_rows", row_ref.get("空例句行", 0)))
                        done_batches += 1
                        mode8_progress.progress(min(1.0, done_batches / max(1, total_batches)))
                        status_text = (
                            f"当前任务：`{row_ref['教材']}`\n\n"
                            f"分配接口：{row_ref['API序号']}\n\n"
                            f"批次进度：{row_ref['已完成']}/{row_ref['总批次']}，"
                            f"新增写入 {row_ref['新增写入']} 条，重复跳过 {row_ref['重复跳过']} 条，空例句 {row_ref['空例句行']} 条。"
                        )
                        if event.get("last_error"):
                            status_text += f"\n\n最近批次异常：{event.get('last_error')}"
                        status_placeholder.info(status_text)
                    elif event_type == "job_done":
                        row_ref["状态"] = "done"
                        row_ref["新增写入"] = int(event.get("inserted_rows", row_ref.get("新增写入", 0)))
                        row_ref["重复跳过"] = int(event.get("skipped_duplicates", row_ref.get("重复跳过", 0)))
                        row_ref["空例句行"] = int(event.get("failed_rows", row_ref.get("空例句行", 0)))
                    render_live()

                if all(f.done() for f in futures):
                    break
                if not had_event:
                    time.sleep(0.08)

            for future, job in futures.items():
                try:
                    summary_rows.append(future.result())
                except Exception as exc:
                    live_rows[job["uid"]]["状态"] = "error"
                    push_debug(
                        job["uid"],
                        {
                            "教材": job["job_label"],
                            "短语": "",
                            "API": live_rows[job["uid"]]["API序号"],
                            "状态": "任务失败",
                            "说明": str(exc),
                        },
                    )
                    st.error(f"`{rel_path(job['excel_file'])}` 处理失败：{exc}")

            render_live()
            st.session_state.mode8_last_summary = summary_rows
            status_placeholder.success("本轮处理已完成，所有任务队列都已执行结束。")

        if st.session_state.mode8_last_summary:
            st.markdown("**最近一次处理结果**")
            render_summary(st.session_state.mode8_last_summary)


if __name__ == "__main__":
    main()
